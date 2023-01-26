import urllib
from _csv import writer

import openpyxl
from bs4 import BeautifulSoup
import os
import csv
import requests
import re
import json
import requests
import pandas as pd
import urllib3
from selenium import webdriver
from time import sleep

inputExcelFile = r"D:\pythonProject\UN_NGO.xlsx"
newWorkbook = openpyxl.load_workbook(inputExcelFile)
file = open(os.path.join(r"D:\pythonProject\ngo_url.csv"), "rU")
reader = csv.reader(file, delimiter=',')
firstWorksheet = newWorkbook['Sheet1']
i = 0
for row in reader:
    for purl in row:

        vgm_url = f'https://esango.un.org/civilsociety/{purl}'
        print(vgm_url)
        html_text = requests.get(vgm_url).text
        soup = BeautifulSoup(html_text, 'html5lib')
        table = soup.find('table', class_="zebraDivided")
        # print(table)
        try:
            on = table.find('td', text="Organization's name:").find_next_sibling("td").text
        except:
            on = ""
        try:
            one = table.find('td', text="Organization's name (English):").find_next_sibling("td").text
        except:
            one = ""
        try:
            oa = table.find('td', text="Organization's acronym:").find_next_sibling("td").text
        except:
            oa = ""
        try:
            oae = table.find('td', text="Organization's acronym (English):").find_next_sibling("td").text
        except:
            oae = ""
        try:
            fn = table.find('td', text="Former Name(s):").find_next_sibling("td").text
        except:
            fn = ""
        try:
            add = table.find('td', text="Address:").find_next_sibling("td").text
        except:
            add = ""
        try:
            ph = table.find('td', text="Phone:").find_next_sibling("td").text
        except:
            ph = ""
        try:
            fax = table.find('td', text="Fax:").find_next_sibling("td").text
        except:
            fax = ""
        try:
            em = table.find('td', text="Email:").find_next_sibling("td").text
        except:
            em = ""
        try:
            web = table.find('td', text="Web site:").find_next_sibling("td").text
        except:
            web = ""
        try:
            ot = table.find('td', text="Organization type:").find_next_sibling("td").text
        except:
            ot = ""
        try:
            lang = table.find('td', text="Languages:").find_next_sibling("td").text
        except:
            lang = ""
        try:
            firstWorksheet.cell(row=i + 1, column=1, value=on)
        except:
            firstWorksheet.cell(row=i + 1, column=1, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=2, value=one)
        except:
            firstWorksheet.cell(row=i + 1, column=2, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=3, value=oa)
        except:
            firstWorksheet.cell(row=i + 1, column=3, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=4, value=oae)
        except:
            firstWorksheet.cell(row=i + 1, column=4, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=5, value=fn)
        except:
            firstWorksheet.cell(row=i + 1, column=5, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=6, value=add)
        except:
            firstWorksheet.cell(row=i + 1, column=6, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=7, value=ph)
        except:
            firstWorksheet.cell(row=i + 1, column=7, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=8, value=fax)
        except:
            firstWorksheet.cell(row=i + 1, column=8, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=9, value=em)
        except:
            firstWorksheet.cell(row=i + 1, column=9, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=10, value=web)
        except:
            firstWorksheet.cell(row=i + 1, column=10, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=11, value=ot)
        except:
            firstWorksheet.cell(row=i + 1, column=11, value="")
        try:
            firstWorksheet.cell(row=i + 1, column=12, value=lang)
        except:
            firstWorksheet.cell(row=i + 1, column=12, value="")
        i += 1
        # print(on, one, oa, oae, fn, add, ph, fax, em, web, ot, lang)
        print(f"{i}: {on}")
        newWorkbook.save(inputExcelFile)